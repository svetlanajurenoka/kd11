from openpyxl import load_workbook

def sixthQuestion():
    count = 0
    for row in ws.iter_rows(2, None, None, None, True):
        if row[3] == "Male":
            count += 1
    return count

def seventhQuestion():
    maxAge = 0
    for row in ws.iter_rows(2, None, None, None, True):
        if row[2] == "IT" and row[4] > maxAge:
            maxAge = row[4]
    return maxAge

def eightQuestion():
    totalAge = 0.0
    count = 0
    for row in ws.iter_rows(2, None, None, None, True):
        if row[2] == "IT":
            totalAge += row[4]
            count += 1
    return round(totalAge/count)

def ninthQuestion():
    count = 0
    for row in ws.iter_rows(2, None, None, None, True):
        if isinstance(row[5], str):
            salary = row[5].replace("$", "").replace(",", ".")
        else:
            salary = row[5]
        if salary > 100000 and salary < 250000:
            count += 1
    return count
        
def tenthQuestion():
    minAge = 0
    # Could make it prettier, but no time :D
    minSalary = 99999999
    for row in ws.iter_rows(2, None, None, None, True):
        if isinstance(row[5], str):
            salary = row[5].replace("$", "").replace(",", ".")
        else:
            salary = row[5] 
        if minSalary > salary:
            minAge = row[4]
    return minAge

wb = load_workbook('EmployeeData.xlsx')
ws = wb.active
total = 0
print("6.jautajums, atbilde: ", sixthQuestion())
print("7.jautajums, atbilde: ", seventhQuestion())
print("8.jautajums, atbilde: ", eightQuestion())
print("9.jautajums, atbilde: ", ninthQuestion())
print("10.jautajums, atbilde: ", tenthQuestion())

# Params for iter_rows: min_row, max_row, min_col, max_col, select_only_values



wb.close()